import pandas as pd
import random
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from generate_timetable import (
    LabTimetable, TheoryTimetable,
    LAB_SUBJECTS, PREBOOKED, DAYS, PERIODS
)
from variables import YEARS_SECTIONS, SUBJECTS_PER_YEAR, TEACHER_ASSIGNMENTS


# ---------------- Helper for borders ----------------
def add_borders(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in [None, ""]:
                cell.border = border


# ---------------- Export Function ----------------
def export_to_excel(filename="timetable_export.xlsx"):
    random.seed(42)
    teacher_busy = defaultdict(lambda: [[False]*PERIODS for _ in range(DAYS)])

    # Generate timetables
    lab_gen = LabTimetable(LAB_SUBJECTS, teacher_busy, prebooked=PREBOOKED)
    lab_gen.generate_lab_pairs()
    theory_gen = TheoryTimetable(
        YEARS_SECTIONS, SUBJECTS_PER_YEAR,
        TEACHER_ASSIGNMENTS, teacher_busy,
        lab_gen.lab_schedules
    )
    theory_gen.generate_timetable()

    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri"]

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 1. Teachers_Subjects - Sorted class wise
        teacher_subjects = []
        for teacher, assignments in TEACHER_ASSIGNMENTS.items():
            for sec, subj in assignments:
                teacher_subjects.append([sec, teacher, subj])
        df_teachers = pd.DataFrame(
            sorted(teacher_subjects, key=lambda x: (x[0], x[2])),
            columns=["Class", "Teacher", "Subject"]
        )
        df_teachers.to_excel(writer, sheet_name="Teachers_Subjects", index=False)

        # 2. Years_Subjects
        year_subjects = []
        for year, subs in SUBJECTS_PER_YEAR.items():
            for subj, freq in subs.items():
                year_subjects.append([year, subj, freq])
        pd.DataFrame(year_subjects, columns=["Year", "Subject", "Frequency"])\
            .to_excel(writer, sheet_name="Years_Subjects", index=False)

        # 3. Lab_Timetables (B1, B2) + borders
        lab_rows = []
        for section in sorted(lab_gen.lab_schedules):
            # B1
            lab_rows.append([f"{section} - B1"])
            lab_rows.append(["Day"] + [f"P{i+1}" for i in range(PERIODS)])
            for d in range(DAYS):
                row = [day_names[d]]
                for p in range(PERIODS):
                    slot = lab_gen.lab_schedules[section]['B1'][d][p]
                    row.append(slot[0] if slot else "-")
                lab_rows.append(row)
            lab_rows.append([])

            # B2
            lab_rows.append([f"{section} - B2"])
            lab_rows.append(["Day"] + [f"P{i+1}" for i in range(PERIODS)])
            for d in range(DAYS):
                row = [day_names[d]]
                for p in range(PERIODS):
                    slot = lab_gen.lab_schedules[section]['B2'][d][p]
                    row.append(slot if slot else "-")
                lab_rows.append(row)
            lab_rows.extend([[], [], [], []])
        pd.DataFrame(lab_rows).to_excel(writer, sheet_name="Lab_Timetables", index=False, header=False)

        # 4. Theory_Timetables + borders
        theory_rows = []
        for section in sorted(theory_gen.class_schedules):
            theory_rows.append([section])
            theory_rows.append(["Day"] + [f"P{i+1}" for i in range(PERIODS)])
            for d in range(DAYS):
                row = [day_names[d]]
                for p in range(PERIODS):
                    slot = theory_gen.class_schedules[section][d][p]
                    row.append(slot if slot else "-")
                theory_rows.append(row)
            theory_rows.extend([[], [], [], []])
        pd.DataFrame(theory_rows).to_excel(writer, sheet_name="Theory_Timetables", index=False, header=False)

        # 5. Merged_Timetables (Lab + Theory)
        merged_rows = []
        for section in sorted(theory_gen.class_schedules):
            merged_rows.append([f"{section} - Merged"])
            merged_rows.append(["Day"] + [f"P{i+1}" for i in range(PERIODS)])
            for d in range(DAYS):
                row = [day_names[d]]
                for p in range(PERIODS):
                    cell_content = "-"
                    if section in lab_gen.lab_schedules and (
                        lab_gen.lab_schedules[section]['B1'][d][p] or lab_gen.lab_schedules[section]['B2'][d][p]
                    ):
                        subjects = []
                        b1 = lab_gen.lab_schedules[section]['B1'][d][p]
                        b2 = lab_gen.lab_schedules[section]['B2'][d][p]
                        if b1:
                            subjects.append(f"{b1[0]}-B1")
                        if b2:
                            subjects.append(f"{b2}-B2")
                        cell_content = "/".join(subjects) if subjects else "-"
                    else:
                        slot = theory_gen.class_schedules[section][d][p]
                        cell_content = slot if slot else "-"
                    row.append(cell_content)
                merged_rows.append(row)
            merged_rows.extend([[], [], [], []])
        pd.DataFrame(merged_rows).to_excel(writer, sheet_name="Merged_Timetables", index=False, header=False)

        # 6. Teacher_Timetables (Subject (Class))
        teacher_tt_rows = []
        for teacher in sorted(TEACHER_ASSIGNMENTS.keys()):
            teacher_tt_rows.append([teacher])
            teacher_tt_rows.append(["Day"] + [f"P{i+1}" for i in range(PERIODS)])
            for d in range(DAYS):
                row = [day_names[d]]
                for p in range(PERIODS):
                    found = "-"
                    # Check labs
                    for section, batches in lab_gen.lab_schedules.items():
                        for batch, sched in batches.items():
                            slot = sched[d][p]
                            if slot and slot[1] == teacher:
                                found = f"{slot} ({section}-{batch})"
                                break
                        if found != "-":
                            break
                    # Check theory
                    if found == "-":
                        for section, sched in theory_gen.class_schedules.items():
                            slot = sched[d][p]
                            if slot and slot[1] == teacher:
                                found = f"{slot} ({section})"
                                break
                    row.append(found)
                teacher_tt_rows.append(row)
            teacher_tt_rows.extend([[], [], [], []])
        pd.DataFrame(teacher_tt_rows).to_excel(writer, sheet_name="Teacher_Timetables", index=False, header=False)

    # Apply borders
    wb = load_workbook(filename)
    for sheet in ["Lab_Timetables", "Theory_Timetables", "Merged_Timetables", "Teacher_Timetables"]:
        add_borders(wb[sheet])
    wb.save(filename)

    print(f"✅ Timetable exported to {filename} with formatting & Pylance-clean code.")


if __name__ == "__main__":
    export_to_excel()
